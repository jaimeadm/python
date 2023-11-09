# pip install openpyxl
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = Workbook()
# worksheet: Worksheet = workbook.active

# Sheet name
sheet_name = 'My Sheet'
# Create sheet
workbook.create_sheet(sheet_name, 0)
# Select sheet
worksheet: Worksheet = workbook[sheet_name]

# Remove a sheet
workbook.remove(workbook['Sheet'])

# Making Header
worksheet.cell(1, 1, 'Name')
worksheet.cell(1, 2, 'Age')
worksheet.cell(1, 3, 'Grade')

students = [
    # name      # age  # grade
    ['João',    14,    5.5],
    ['Maria',   13,    9.7],
    ['Luiz',    15,    8.8],
    ['Claudio', 16,    10],
]

# for i, student_row in enumerate(students, start=2):
#     for j, student_column in enumerate(student_row, start=1):
#         worksheet.cell(i, j, student_column)

for student in students:
    worksheet.append(student)

workbook.save(WORKBOOK_PATH)
